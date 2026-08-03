"""大会運営ガイド（docs/operation-guide.md）を Excel ブックに書き出す。

    python tools/build_operation_xlsx.py

**正本は Markdown のほうです。** 内容を直すときは `docs/operation-guide.md` を編集し、
このスクリプトを流し直してください。Excel を直接編集しても、次の実行で消えます。

`##` の見出しごとにシートを分け、当日そのまま印刷して手元に置ける体裁にしています。
配色はサイトと同じ桃色・金です。
"""
import os
import re
import sys
import unicodedata

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
SRC = os.path.join(ROOT, "docs", "operation-guide.md")
OUT = os.path.join(ROOT, "docs", "大会運営ガイド.xlsx")

BODY = "Meiryo"                     # 日本語が化けない標準フォント
MONO = "Consolas"

# サイトと同じ色（assets/style.css の :root より）
PINK = "D22C6B"
PINK_SOFT = "FFEDF2"
GOLD_PALE = "FFF3D6"
INK = "3A2430"
GREY = "F2F2F2"
WHITE = "FFFFFF"

COLS = 4                            # A〜D。表が最大3列なので4列あれば足りる
WIDTHS = [30, 34, 44, 16]

thin = Side(style="thin", color="E6D2DA")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)


# ---------- Markdown を読む ----------

def strip_marks(text):
    """セルに入れる前に、Markdown の記号だけ落とす（太字や `code` の中身は残す）。"""
    text = re.sub(r"\*\*(.+?)\*\*", r"\1", text)
    text = re.sub(r"`(.+?)`", r"\1", text)
    text = re.sub(r"\[(.+?)\]\((.+?)\)", r"\1", text)
    return text.strip()


def is_emphasized(text):
    """行の中に ** があれば「強調したい行」とみなす。"""
    return "**" in text


def parse(md):
    """Markdown を [(シート名, [ブロック…])] にする。

    ブロックは ("h3"|"p"|"ul"|"code"|"quote"|"table"|"check", 中身) の形。
    """
    sheets, blocks = [], []
    name = "はじめに"
    lines = md.split("\n")
    i = 0
    while i < len(lines):
        line = lines[i]
        before = i                                      # 進まなかったときの保険（下の末尾で使う）

        if re.match(r"^-{3,}\s*$", line):               # 水平線。捨てる
            i += 1                                      # ※ここを落とすと「-」始まりなのに
            continue                                    #   箇条書きにも段落にもならず止まります
        if line.startswith("## "):
            sheets.append((name, blocks))
            name, blocks = sheet_name(line[3:]), []
            i += 1
        elif line.startswith("# "):
            i += 1
        elif line.startswith("### "):
            blocks.append(("h3", strip_marks(line[4:])))
            i += 1
        elif line.startswith("```"):
            buf, i = [], i + 1
            while i < len(lines) and not lines[i].startswith("```"):
                buf.append(lines[i])
                i += 1
            i += 1
            blocks.append(("code", "\n".join(buf).rstrip()))
        elif line.startswith("|"):
            rows = []
            while i < len(lines) and lines[i].startswith("|"):
                cells = [c.strip() for c in lines[i].strip().strip("|").split("|")]
                if not all(set(c) <= set("-: ") for c in cells):   # 区切り行は捨てる
                    rows.append([strip_marks(c) for c in cells])
                i += 1
            blocks.append(("table", rows))
        elif line.startswith("> "):
            buf = []
            while i < len(lines) and (lines[i].startswith(">") or
                                      (buf and lines[i].strip() and not lines[i].startswith(("#", "|", "-", "```")))):
                buf.append(lines[i].lstrip("> ").rstrip())
                i += 1
            blocks.append(("quote", strip_marks(" ".join(b for b in buf if b))))
        elif re.match(r"^- \[ \] ", line):
            items = []
            while i < len(lines) and re.match(r"^- \[ \] ", lines[i]):
                items.append(strip_marks(lines[i][6:]))
                i += 1
            blocks.append(("check", items))
        elif line.startswith("- "):
            items = []
            while i < len(lines) and (lines[i].startswith("- ") or lines[i].startswith("  ")):
                if lines[i].startswith("- "):
                    items.append([strip_marks(lines[i][2:]), is_emphasized(lines[i])])
                else:                                              # 折り返しの続き
                    items[-1][0] += " " + strip_marks(lines[i])
                i += 1
            blocks.append(("ul", items))
        elif line.strip():
            buf = []
            while i < len(lines) and lines[i].strip() and not lines[i].startswith(("#", "|", "-", ">", "```")):
                buf.append(lines[i].strip())
                i += 1
            joined = " ".join(buf)
            blocks.append(("p", (strip_marks(joined), is_emphasized(joined))))
        else:
            i += 1

        # どの分岐でも行が進まなかったら、1行送って抜ける。
        # 想定外の記法が来ても止まらないようにするための保険です。
        if i == before:
            i += 1

    sheets.append((name, blocks))
    return [(n, b) for n, b in sheets if b]


def sheet_name(title):
    """「1. 応募が締め切られたら（前日）」→「1 前日」のように短くする。"""
    short = {
        "0": "0 事前準備", "1": "1 前日", "2": "2 当日",
        "3": "3 飛び入り", "4": "4 困ったとき", "5": "5 当日チェック",
    }
    m = re.match(r"^(\d+)\.", title.strip())
    if m and m.group(1) in short:
        return short[m.group(1)]
    return re.sub(r"[\[\]:*?/\\]", "", strip_marks(title))[:31]


# ---------- Excel に書く ----------

# 1行に入る幅。列幅の合計（124）から、字下げと余白のぶんを引いた値。
# 単位は半角1文字ぶんなので、日本語は2つ使う。
WIDTH_BODY = 112
WIDTH_CODE = 125                    # コードは Consolas 10pt で本文より小さいぶん多く入る


def display_width(text):
    """全角を2、半角を1として数えた表示幅。"""
    return sum(2 if unicodedata.east_asian_width(ch) in "WFA" else 1 for ch in str(text))


def wrapped_height(text, size=11, per_line=WIDTH_BODY, line_h=None):
    """折り返したときに必要な行の高さを見積もる。

    **Excel は、結合したセルの中身が折り返しても行の高さを自動調整しません。**
    指定しないまま長文を入れると下が切れて読めなくなるため、ここで見積もっています。
    日本語と英数字では1行に入る量が倍違うので、文字数ではなく表示幅で数えます。
    """
    line_h = line_h or (size + 4)
    lines = 0
    for para in str(text).split("\n"):
        lines += max(1, -(-display_width(para) // per_line))  # 切り上げ
    return lines * line_h + 4


def merge_row(ws, r, text, *, font=None, fill=None, align=None, indent=0, height=None):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
    c = ws.cell(row=r, column=1, value=text)
    c.font = font or Font(name=BODY, size=11, color=INK)
    c.alignment = align or Alignment(wrap_text=True, vertical="top", indent=indent)
    if fill:
        for col in range(1, COLS + 1):
            ws.cell(row=r, column=col).fill = fill
    ws.row_dimensions[r].height = height if height else wrapped_height(text)
    return r + 1


def write_sheet(wb, name, blocks, index):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    r = merge_row(ws, r, name,
                  font=Font(name=BODY, size=16, bold=True, color=WHITE),
                  fill=PatternFill("solid", fgColor=PINK),
                  align=Alignment(vertical="center", indent=1), height=30)
    r += 1

    for kind, body in blocks:
        if kind == "h3":
            r = merge_row(ws, r, body,
                          font=Font(name=BODY, size=12, bold=True, color=PINK),
                          fill=PatternFill("solid", fgColor=PINK_SOFT),
                          align=Alignment(vertical="center", indent=1), height=24)

        elif kind == "p":
            text, strong = body
            r = merge_row(ws, r, text,
                          font=Font(name=BODY, size=11, bold=strong, color=INK), indent=1)

        elif kind == "ul":
            for text, strong in body:
                r = merge_row(ws, r, "・" + text,
                              font=Font(name=BODY, size=11, bold=strong, color=INK), indent=1)

        elif kind == "check":
            for text in body:
                ws.cell(row=r, column=1, value="☐").font = Font(name=BODY, size=14, color=PINK)
                ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=r, column=1).border = BOX
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=COLS)
                c = ws.cell(row=r, column=2, value=text)
                c.font = Font(name=BODY, size=11, color=INK)
                c.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
                for col in range(2, COLS + 1):
                    ws.cell(row=r, column=col).border = BOX
                ws.row_dimensions[r].height = 26
                r += 1

        elif kind == "code":
            # 実行結果を貼っている箇所は1行が長く、折り返すので幅で数える
            r = merge_row(ws, r, body,
                          font=Font(name=MONO, size=10, color=INK),
                          fill=PatternFill("solid", fgColor=GREY),
                          align=Alignment(wrap_text=True, vertical="top", indent=1),
                          height=wrapped_height(body, per_line=WIDTH_CODE, line_h=14))

        elif kind == "quote":
            r = merge_row(ws, r, "！ " + body,
                          font=Font(name=BODY, size=11, bold=True, color=INK),
                          fill=PatternFill("solid", fgColor=GOLD_PALE), indent=1)

        elif kind == "table":
            head, *rest = body
            for j, cell in enumerate(head[:COLS], start=1):
                c = ws.cell(row=r, column=j, value=cell)
                c.font = Font(name=BODY, size=10, bold=True, color=WHITE)
                c.fill = PatternFill("solid", fgColor=PINK)
                c.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
                c.border = BOX
            ws.row_dimensions[r].height = 22
            r += 1
            for row in rest:
                for j, cell in enumerate(row[:COLS], start=1):
                    c = ws.cell(row=r, column=j, value=cell)
                    c.font = Font(name=BODY, size=10, color=INK)
                    c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
                    c.border = BOX
                # 表は結合していないので Excel が自動調整するが、いちばん長い
                # セルに合わせて下限を決めておくと、印刷したときに詰まらない
                longest = max((len(str(c or "")) for c in row[:COLS]), default=0)
                col_chars = max(8, WIDTHS[0] // 2)
                ws.row_dimensions[r].height = max(18, -(-longest // col_chars) * 15 + 4)
                r += 1

        r += 1                                          # ブロックの間に1行あける

    # 印刷して手元に置く前提。1ページ幅に収め、見出し行は各ページに繰り返す
    ws.print_title_rows = "1:1"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.freeze_panes = "A2"
    return ws


def build_index(wb, sheets):
    ws = wb.create_sheet("目次", 0)
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([30, 52, 30, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:D1")
    c = ws.cell(row=1, column=1, value="ねっ子ポカジャン大会　運営ガイド")
    c.font = Font(name=BODY, size=18, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=PINK)
    c.alignment = Alignment(vertical="center", indent=1)
    for col in range(1, 5):
        ws.cell(row=1, column=col).fill = PatternFill("solid", fgColor=PINK)
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:D2")
    c = ws.cell(row=2, column=1,
                value="当日はこのブックを開いたまま進めてください。左下のタブが手順の順番です。")
    c.font = Font(name=BODY, size=11, color=INK)
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[2].height = 22

    hint = {
        "はじめに": "このガイドの読みかたと、関連する文書",
        "0 事前準備": "当日までに終わらせておくこと。未設定の項目がここに出ています",
        "1 前日": "参加者CSVの用意 → 抽選 → 進行の確認",
        "2 当日": "卓分けの発表 → 結果の入力 → サイトへ反映",
        "3 飛び入り": "人数が変わったときの組み直しかた",
        "4 困ったとき": "エラーが出たときの読み解きかたと対処",
        "5 当日チェック": "始める前に確認する持ち物リスト",
    }
    r = 4
    for j, title in enumerate(["シート", "内容"], start=1):
        c = ws.cell(row=r, column=j, value=title)
        c.font = Font(name=BODY, size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=PINK)
        c.alignment = Alignment(vertical="center", indent=1)
        c.border = BOX
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    for col in range(2, 5):
        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=PINK)
        ws.cell(row=r, column=col).border = BOX
    r += 1

    for name in sheets:
        c = ws.cell(row=r, column=1, value=name)
        c.font = Font(name=BODY, size=11, bold=True, color=PINK)
        c.alignment = Alignment(vertical="center", indent=1)
        c.border = BOX
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        c2 = ws.cell(row=r, column=2, value=hint.get(name, ""))
        c2.font = Font(name=BODY, size=10, color=INK)
        c2.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
        for col in range(2, 5):
            ws.cell(row=r, column=col).border = BOX
        ws.row_dimensions[r].height = 24
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    c = ws.cell(row=r, column=1,
                value="！ このブックは docs/operation-guide.md から作っています。"
                      "内容を直すときは Markdown のほうを編集し、"
                      "python tools/build_operation_xlsx.py を流し直してください。"
                      "Excel を直接編集しても次の実行で消えます。")
    c.font = Font(name=BODY, size=10, bold=True, color=INK)
    c.fill = PatternFill("solid", fgColor=GOLD_PALE)
    c.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
    for col in range(1, 5):
        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=GOLD_PALE)
    ws.row_dimensions[r].height = 46


def main():
    if not os.path.exists(SRC):
        sys.exit(f"元の文書が見つかりません: {SRC}")

    with open(SRC, encoding="utf-8") as f:
        sheets = parse(f.read())

    wb = Workbook()
    wb.remove(wb.active)
    for i, (name, blocks) in enumerate(sheets):
        write_sheet(wb, name, blocks, i)
    build_index(wb, [n for n, _ in sheets])
    wb.active = 0
    wb.save(OUT)

    print(f"書き出しました: {OUT}")
    for name, blocks in sheets:
        print(f"  {name}（{len(blocks)}ブロック）")


if __name__ == "__main__":
    main()
